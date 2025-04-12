import openpyxl
book = openpyxl.load_workbook("Book1.xlsx")
sheet = book.active
# print(sheet.cell(1,1).value)
# sheet.cell(2,2).value = 'Kartik'
# sheet.cell(3,2).value = 'Shinde'
# print(sheet['A5'].value)

rows = sheet.max_row
columns = sheet.max_column
dict = {}
for i in range(1, rows+1):
    if sheet.cell(i, 1).value == 'TestCase1':
        for j in range(2, columns+1):
            dict.update({sheet.cell(1,j).value : sheet.cell(i, j).value})
print(dict)